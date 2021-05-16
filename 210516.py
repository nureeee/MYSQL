import openpyxl as xl

exf = xl.load_workbook('c:\\dd\\itsal.xlsx')

aws = exf.active

tot = 0

for i in aws.rows:
    index = i[0].rows
    name = i[0].value
    salary = i[1].value

    tot += salary
    last_rows = len(tuple(aws.rows))+1
    aws.cell(row = last_rows, colum = 1).value = '평균'
    aws.cell(row = last_rows, colum = 2).value = avg

    print('{} {}'.format(name, salary))


    exf.save('c:\\dd\\outits.xlsx')
    exf.close()